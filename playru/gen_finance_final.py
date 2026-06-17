from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

F = {
    "header": PatternFill("solid", fgColor="1E1B4B"),
    "green": PatternFill("solid", fgColor="D1FAE5"),
    "red": PatternFill("solid", fgColor="FEE2E2"),
    "yellow": PatternFill("solid", fgColor="FEF3C7"),
    "purple": PatternFill("solid", fgColor="EEF2FF"),
    "input": PatternFill("solid", fgColor="E0E7FF"),
}
F_BOLD = Font(name="Segoe UI", bold=True, size=10, color="1E1B4B")
F_NORM = Font(name="Segoe UI", size=10, color="1E1B4B")
F_HDR = Font(name="Segoe UI", bold=True, color="FFFFFF", size=11)
F_SEC = Font(name="Segoe UI", bold=True, size=10, color="1E1B4B")
F_BIG = Font(name="Segoe UI", size=14, bold=True, color="7C3AED")
F_INP = Font(name="Segoe UI", size=11, color="1E1B4B", bold=True)
THIN = Border(left=Side(style="thin", color="E2E4EA"), right=Side(style="thin", color="E2E4EA"),
              top=Side(style="thin", color="E2E4EA"), bottom=Side(style="thin", color="E2E4EA"))

MONTHS = "Январь|Февраль|Март|Апрель|Май|Июнь|Июль|Август|Сентябрь|Октябрь|Ноябрь|Декабрь".split("|")
MC, TOT = 14, 14

def hdr(r):
    for c in range(1, MC+1):
        cl = ws.cell(row=r, column=c)
        cl.fill = F["header"]; cl.font = F_HDR; cl.border = THIN
        cl.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def sv(r, fk=None):
    ff = F.get(fk)
    for c in range(1, MC+1):
        cl = ws.cell(row=r, column=c)
        if ff: cl.fill = ff
        cl.font = F_NORM; cl.border = THIN
        cl.alignment = Alignment(horizontal="center" if c > 1 else "left", vertical="center")
        if c > 1: cl.number_format = '#,##0.00'

def sec(r, text, fk=None):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=MC)
    cl = ws.cell(row=r, column=1, value=text)
    if fk: cl.fill = F[fk]
    cl.font = F_SEC; cl.border = THIN
    for c in range(2, MC+1): ws.cell(row=r, column=c).border = THIN

def sv2(r, fk=None):
    ff = F.get(fk)
    for c in range(1, 6):
        cl = ws2.cell(row=r, column=c)
        if ff: cl.fill = ff
        cl.font = F_NORM; cl.border = THIN
        cl.alignment = Alignment(horizontal="center" if c > 1 else "left", vertical="center")

def sec2(r, text, fk=None):
    ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    cl = ws2.cell(row=r, column=1, value=text)
    if fk: cl.fill = F[fk]
    cl.font = F_SEC; cl.border = THIN
    for c in range(2, 6): ws2.cell(row=r, column=c).border = THIN

# ════════════ ЛИСТ 1: ГОД ════════════
ws = wb.active
ws.title = "Год"
ws.column_dimensions["A"].width = 36
for i in range(2, MC+1): ws.column_dimensions[get_column_letter(i)].width = 13

ws.cell(row=1, column=1, value="Категория")
for i, m in enumerate(MONTHS, 2): ws.cell(row=1, column=i, value=m)
ws.cell(row=1, column=TOT, value="Итого за год")
hdr(1)
r = 2

# ── ДОХОДЫ ──  (ТЕПЕРЬ 2 СТРОКИ ЗАРПЛАТЫ)
sec(r, "ДОХОДЫ", "green"); r += 1
ic = ["Зарплата (аванс)", "Зарплата (основная)", "Инвестиции", "Прочие доходы"]
ir1 = r
for cat in ic:
    ws.cell(row=r, column=1, value=cat)
    for m in range(2, TOT+1):
        ws.cell(row=r, column=m, value=0 if m < TOT else f"=SUM(B{r}:M{r})")
    sv(r); r += 1
ir2 = r - 1
ws.cell(row=r, column=1, value="Итого доходов")
for m in range(2, TOT+1):
    cl = get_column_letter(m)
    ws.cell(row=r, column=m, value=f"=SUM({cl}{ir1}:{cl}{ir2})")
sv(r, "green"); tir = r; r += 1

# ── РАСХОДЫ ──
sec(r, "РАСХОДЫ", "red"); r += 1
ec = ["Продукты","Коммунальные","Транспорт","Связь/Интернет",
      "Развлечения","Одежда","Здоровье","Рестораны","Прочее"]
er1 = r
for cat in ec:
    ws.cell(row=r, column=1, value=cat)
    for m in range(2, TOT+1):
        ws.cell(row=r, column=m, value=0 if m < TOT else f"=SUM(B{r}:M{r})")
    sv(r); r += 1
er2 = r - 1
ws.cell(row=r, column=1, value="Итого расходов")
for m in range(2, TOT+1):
    cl = get_column_letter(m)
    ws.cell(row=r, column=m, value=f"=SUM({cl}{er1}:{cl}{er2})")
sv(r, "red"); ter = r; r += 1

# ── КРЕДИТЫ (кредитные карты) ──
sec(r, "КРЕДИТЫ (кредитные карты)", "yellow"); r += 1

# Сбербанк
ws.cell(row=r, column=1, value="СБЕРБАНК (кредитная карта)").font = F_BOLD
for m in range(2, MC+1): ws.cell(row=r, column=m).border = THIN
r += 1

sb_lim = r; sb_debt = r+1; sb_rate = r+2; sb_pay = r+3
sb_int = r+4; sb_payd = r+5; sb_end = r+6

ws.cell(row=sb_lim, column=1, value="  Кредитный лимит")
for m in range(2, TOT+1):
    if m < TOT: ws.cell(row=sb_lim, column=m, value=300000)
    else: ws.cell(row=sb_lim, column=m, value=f"={get_column_letter(m-1)}{sb_lim}")
ws.cell(row=sb_lim, column=2).fill = F["input"]; sv(sb_lim); r = sb_lim+1

ws.cell(row=sb_debt, column=1, value="  Задолженность на начало")
for m in range(2, TOT+1):
    cl = get_column_letter(m)
    if m == 2:
        ws.cell(row=sb_debt, column=m, value=150000)
        ws.cell(row=sb_debt, column=m).fill = F["input"]
    elif m < TOT:
        ws.cell(row=sb_debt, column=m, value=f"={get_column_letter(m-1)}{sb_end}")
    else:
        ws.cell(row=sb_debt, column=m, value=f"=SUM(B{sb_debt}:M{sb_debt})")
sv(sb_debt); r = sb_debt+1

ws.cell(row=sb_rate, column=1, value="  Ставка % годовых")
ws.cell(row=sb_rate, column=2, value=25).fill = F["input"]
ws.cell(row=sb_rate, column=2).number_format = '0.0'
for m in range(3, TOT+1):
    ws.cell(row=sb_rate, column=m, value=f"=B{sb_rate}")
    ws.cell(row=sb_rate, column=m).number_format = '0.0'
sv(sb_rate); r = sb_rate+1

ws.cell(row=sb_pay, column=1, value="  Обязательный взнос")
for m in range(2, TOT+1):
    if m < TOT: ws.cell(row=sb_pay, column=m, value=7000)
    else: ws.cell(row=sb_pay, column=m, value=f"=SUM(B{sb_pay}:M{sb_pay})")
ws.cell(row=sb_pay, column=2).fill = F["input"]; sv(sb_pay); r = sb_pay+1

ws.cell(row=sb_int, column=1, value="  Начислено процентов")
for m in range(2, TOT+1):
    cl = get_column_letter(m)
    ws.cell(row=sb_int, column=m).value = f"={cl}{sb_debt}*{cl}{sb_rate}/12/100"
    if m == TOT: ws.cell(row=sb_int, column=m).value = f"=SUM(B{sb_int}:M{sb_int})"
sv(sb_int); r = sb_int+1

ws.cell(row=sb_payd, column=1, value="  Погашение долга")
for m in range(2, TOT+1):
    cl = get_column_letter(m)
    ws.cell(row=sb_payd, column=m).value = f"=IF({cl}{sb_pay}-{cl}{sb_int}>0,{cl}{sb_pay}-{cl}{sb_int},0)"
    if m == TOT: ws.cell(row=sb_payd, column=m).value = f"=SUM(B{sb_payd}:M{sb_payd})"
sv(sb_payd); r = sb_payd+1

ws.cell(row=sb_end, column=1, value="  Остаток на конец")
for m in range(2, TOT+1):
    cl = get_column_letter(m)
    ws.cell(row=sb_end, column=m).value = f"={cl}{sb_debt}-{cl}{sb_payd}"
    if m == TOT: ws.cell(row=sb_end, column=m).value = f"=M{sb_end}"
sv(sb_end); r = sb_end+2

# Т-Банк
ws.cell(row=r, column=1, value="Т-БАНК (кредитная карта)").font = F_BOLD
for m in range(2, MC+1): ws.cell(row=r, column=m).border = THIN
r += 1

tt_lim = r; tt_debt = r+1; tt_rate = r+2; tt_pay = r+3
tt_int = r+4; tt_payd = r+5; tt_end = r+6

ws.cell(row=tt_lim, column=1, value="  Кредитный лимит")
for m in range(2, TOT+1):
    if m < TOT: ws.cell(row=tt_lim, column=m, value=200000)
    else: ws.cell(row=tt_lim, column=m, value=f"={get_column_letter(m-1)}{tt_lim}")
ws.cell(row=tt_lim, column=2).fill = F["input"]; sv(tt_lim); r = tt_lim+1

ws.cell(row=tt_debt, column=1, value="  Задолженность на начало")
for m in range(2, TOT+1):
    cl = get_column_letter(m)
    if m == 2:
        ws.cell(row=tt_debt, column=m, value=50000)
        ws.cell(row=tt_debt, column=m).fill = F["input"]
    elif m < TOT:
        ws.cell(row=tt_debt, column=m, value=f"={get_column_letter(m-1)}{tt_end}")
    else:
        ws.cell(row=tt_debt, column=m, value=f"=SUM(B{tt_debt}:M{tt_debt})")
sv(tt_debt); r = tt_debt+1

ws.cell(row=tt_rate, column=1, value="  Ставка % годовых")
ws.cell(row=tt_rate, column=2, value=28).fill = F["input"]
ws.cell(row=tt_rate, column=2).number_format = '0.0'
for m in range(3, TOT+1):
    ws.cell(row=tt_rate, column=m, value=f"=B{tt_rate}")
    ws.cell(row=tt_rate, column=m).number_format = '0.0'
sv(tt_rate); r = tt_rate+1

ws.cell(row=tt_pay, column=1, value="  Обязательный взнос")
for m in range(2, TOT+1):
    if m < TOT: ws.cell(row=tt_pay, column=m, value=5000)
    else: ws.cell(row=tt_pay, column=m, value=f"=SUM(B{tt_pay}:M{tt_pay})")
ws.cell(row=tt_pay, column=2).fill = F["input"]; sv(tt_pay); r = tt_pay+1

ws.cell(row=tt_int, column=1, value="  Начислено процентов")
for m in range(2, TOT+1):
    cl = get_column_letter(m)
    ws.cell(row=tt_int, column=m).value = f"={cl}{tt_debt}*{cl}{tt_rate}/12/100"
    if m == TOT: ws.cell(row=tt_int, column=m).value = f"=SUM(B{tt_int}:M{tt_int})"
sv(tt_int); r = tt_int+1

ws.cell(row=tt_payd, column=1, value="  Погашение долга")
for m in range(2, TOT+1):
    cl = get_column_letter(m)
    ws.cell(row=tt_payd, column=m).value = f"=IF({cl}{tt_pay}-{cl}{tt_int}>0,{cl}{tt_pay}-{cl}{tt_int},0)"
    if m == TOT: ws.cell(row=tt_payd, column=m).value = f"=SUM(B{tt_payd}:M{tt_payd})"
sv(tt_payd); r = tt_payd+1

ws.cell(row=tt_end, column=1, value="  Остаток на конец")
for m in range(2, TOT+1):
    cl = get_column_letter(m)
    ws.cell(row=tt_end, column=m).value = f"={cl}{tt_debt}-{cl}{tt_payd}"
    if m == TOT: ws.cell(row=tt_end, column=m).value = f"=M{tt_end}"
sv(tt_end); r = tt_end+2

tcr = r
ws.cell(row=r, column=1, value="Итого кредитные взносы")
for m in range(2, TOT+1):
    cl = get_column_letter(m)
    ws.cell(row=r, column=m).value = f"={cl}{sb_pay}+{cl}{tt_pay}"
sv(r, "yellow"); r += 2

# ── ИТОГИ ──
ws.cell(row=r, column=1, value="Остаток (Доходы - Расходы - Кредиты)")
for m in range(2, TOT+1):
    cl = get_column_letter(m)
    ws.cell(row=r, column=m).value = f"={cl}{tir}-{cl}{ter}-{cl}{tcr}"
sv(r); nor = r; r += 1

ws.cell(row=r, column=1, value="Накоплено нарастающим итогом")
for m in range(2, TOT+1):
    cl = get_column_letter(m)
    if m == 2: ws.cell(row=r, column=m, value=f"={cl}{nor}")
    else:
        pv = get_column_letter(m-1)
        ws.cell(row=r, column=m, value=f"={pv}{r}+{cl}{nor}")
sv(r); r += 1

ws.freeze_panes = "B2"

# ════════════ ЛИСТ 2: СВОДКА ════════════
ws2 = wb.create_sheet("Сводка")
ws2.column_dimensions["A"].width = 34
ws2.column_dimensions["B"].width = 18
ws2.column_dimensions["C"].width = 18
ws2.column_dimensions["D"].width = 12
ws2.column_dimensions["E"].width = 38

for i, h in enumerate(["Категория","Сумма за год","Среднее в месяц","Доля","Примечание"], 1):
    cl = ws2.cell(row=1, column=i, value=h)
    cl.fill = F["header"]; cl.font = F_HDR; cl.border = THIN
    cl.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

SR = 2
sec2(SR, "ДОХОДЫ", "green"); SR += 1
for cat in ic:
    ws2.cell(row=SR, column=1, value=cat)
    ws2.cell(row=SR, column=2, value=f"='Год'!N{ir1}")
    ws2.cell(row=SR, column=3, value=f"=B{SR}/12")
    sv2(SR); ir1 += 1; SR += 1
ws2.cell(row=SR, column=1, value="Итого доходов")
ws2.cell(row=SR, column=2, value=f"='Год'!N{tir}")
ws2.cell(row=SR, column=3, value=f"=B{SR}/12")
ws2.cell(row=SR, column=4, value="100%")
sv2(SR, "green"); SR += 2

sec2(SR, "РАСХОДЫ", "red"); SR += 1
for cat in ec:
    ws2.cell(row=SR, column=1, value=cat)
    ws2.cell(row=SR, column=2, value=f"='Год'!N{er1}")
    ws2.cell(row=SR, column=3, value=f"=B{SR}/12")
    sv2(SR); er1 += 1; SR += 1
ws2.cell(row=SR, column=1, value="Итого расходов")
ws2.cell(row=SR, column=2, value=f"='Год'!N{ter}")
ws2.cell(row=SR, column=3, value=f"=B{SR}/12")
ws2.cell(row=SR, column=4, value="100%")
sv2(SR, "red"); SR += 2

sec2(SR, "КРЕДИТЫ", "yellow"); SR += 1
ws2.cell(row=SR, column=1, value="Сбербанк — взносы за год")
ws2.cell(row=SR, column=2, value=f"='Год'!N{sb_pay}"); sv2(SR); SR += 1
ws2.cell(row=SR, column=1, value="Т-Банк — взносы за год")
ws2.cell(row=SR, column=2, value=f"='Год'!N{tt_pay}"); sv2(SR); SR += 1
ws2.cell(row=SR, column=1, value="Итого кредитные взносы")
ws2.cell(row=SR, column=2, value=f"='Год'!N{tcr}"); sv2(SR, "yellow"); SR += 2

sec2(SR, "ЗАДОЛЖЕННОСТЬ ПО КРЕДИТАМ", "purple"); SR += 1
ws2.cell(row=SR, column=1, value="Сбербанк — долг на конец года")
ws2.cell(row=SR, column=2, value=f"='Год'!N{sb_end}"); sv2(SR); SR += 1
ws2.cell(row=SR, column=1, value="Т-Банк — долг на конец года")
ws2.cell(row=SR, column=2, value=f"='Год'!N{tt_end}"); sv2(SR); SR += 1
ws2.cell(row=SR, column=1, value="Общая задолженность")
ws2.cell(row=SR, column=2, value=f"=B{SR-2}+B{SR-1}"); sv2(SR, "purple"); SR += 2

sec2(SR, "ФИНАНСОВЫЙ РЕЗУЛЬТАТ", "header"); SR += 1
ws2.cell(row=SR, column=1, value="Всего доходов")
ws2.cell(row=SR, column=2, value=f"='Год'!N{tir}"); sv2(SR); inc_r = SR; SR += 1
ws2.cell(row=SR, column=1, value="Всего расходов")
ws2.cell(row=SR, column=2, value=f"='Год'!N{ter}"); sv2(SR); exp_r = SR; SR += 1
ws2.cell(row=SR, column=1, value="Всего кредитных взносов")
ws2.cell(row=SR, column=2, value=f"='Год'!N{tcr}"); sv2(SR); cre_r = SR; SR += 1
ws2.cell(row=SR, column=1, value="Чистый остаток за год")
ws2.cell(row=SR, column=2, value=f"=B{inc_r}-B{exp_r}-B{cre_r}")
ws2.cell(row=SR, column=3, value=f"=B{SR}/12")
ws2.cell(row=SR, column=5, value="Отрицательный = перерасход")
sv2(SR); SR += 2

ws2.cell(row=SR, column=1, value="Коэффициент сбережения")
ws2.cell(row=SR, column=2, value=f"=IF(B{inc_r}>0,(B{inc_r}-B{exp_r}-B{cre_r})/B{inc_r},0)")
ws2.cell(row=SR, column=2).number_format = '0.0%'
ws2.cell(row=SR, column=5, value=">10% — здоровый бюджет"); sv2(SR); SR += 1
ws2.cell(row=SR, column=1, value="Доля кредитных взносов")
ws2.cell(row=SR, column=2, value=f"=IF(B{exp_r}>0,B{cre_r}/B{exp_r},0)")
ws2.cell(row=SR, column=2).number_format = '0.0%'
ws2.cell(row=SR, column=5, value="Рекомендуется <30%"); sv2(SR); SR += 1
ws2.freeze_panes = "B2"

# ════════════ ЛИСТ 3: НАКОПЛЕНИЯ ════════════
ws3 = wb.create_sheet("Накопления")
for c, w in enumerate([8, 22, 22, 22, 22, 26, 18], 1):
    ws3.column_dimensions[get_column_letter(c)].width = w

ws3.merge_cells("A1:G1")
cl = ws3.cell(row=1, column=1, value="КАЛЬКУЛЯТОР НАКОПЛЕНИЙ")
cl.fill = F["header"]; cl.font = F_HDR
cl.alignment = Alignment(horizontal="center", vertical="center")
for c in range(1, 8): ws3.cell(row=1, column=c).border = THIN

for ro, lbl, val, fmt in [
    (2, "Цель:", 300000, '#,##0.00'),
    (3, "Взнос в месяц:", 15000, '#,##0.00'),
    (4, "Годовая ставка:", 15, '0.0'),
]:
    ws3.cell(row=ro, column=1, value=lbl).font = F_BOLD
    ws3.merge_cells(start_row=ro, start_column=1, end_row=ro, end_column=2)
    ws3.cell(row=ro, column=3, value=val).fill = F["input"]
    ws3.cell(row=ro, column=3).font = F_INP; ws3.cell(row=ro, column=3).number_format = fmt
    ws3.cell(row=ro, column=3).border = THIN
    unit = "руб." if ro != 4 else "%"
    ws3.cell(row=ro, column=4, value=unit).border = THIN
    for c in [1, 2, 5, 6, 7]: ws3.cell(row=ro, column=c).border = THIN

MAX_R = 120
hr = 6
for i, h in enumerate(["Месяц","Взнос","Остаток на начало","Проценты","Взнос + %","Итого накоплено","До цели"], 1):
    cl = ws3.cell(row=hr, column=i, value=h)
    cl.fill = F["header"]; cl.font = F_HDR; cl.border = THIN
    cl.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

for i in range(MAX_R):
    rr = hr + 1 + i
    ws3.cell(row=rr, column=1, value=i + 1)
    ws3.cell(row=rr, column=2, value="$C$3")
    ws3.cell(row=rr, column=3, value=0 if i == 0 else f"=F{rr-1}")
    ws3.cell(row=rr, column=4, value=f"=C{rr}*($C$4/100/12)")
    ws3.cell(row=rr, column=5, value=f"=B{rr}+D{rr}")
    ws3.cell(row=rr, column=6, value=f"=C{rr}+E{rr}")
    ws3.cell(row=rr, column=7, value=f"=IF(F{rr}>=$C$2,0,$C$2-F{rr})")
    for c in range(1, 8):
        cl = ws3.cell(row=rr, column=c)
        cl.font = F_NORM; cl.border = THIN
        cl.alignment = Alignment(horizontal="center" if c > 1 else "left", vertical="center")
        if c > 1: cl.number_format = '#,##0.00'

# Результат
res = hr + 1 + MAX_R + 1
ws3.merge_cells(start_row=res, start_column=1, end_row=res, end_column=7)
cl = ws3.cell(row=res, column=1, value="РЕЗУЛЬТАТ")
cl.fill = F["header"]; cl.font = F_HDR
cl.alignment = Alignment(horizontal="center", vertical="center")
for c in range(1, 8): ws3.cell(row=res, column=c).border = THIN
res += 1

# Строка: сколько месяцев нужно (используем MATCH + INDEX)
# F{hr+1}:F{hr+MAX_R} — диапазон накоплений, ищем когда >= цели
ws3.cell(row=res, column=1, value="Месяцев до цели:").font = F_BOLD
ws3.merge_cells(start_row=res, start_column=1, end_row=res, end_column=3)
match_formula = f'=IFERROR(MATCH(TRUE,INDEX(F{hr+1}:F{hr+MAX_R}>=$C$2,0),0),"более {MAX_R}")'
ws3.cell(row=res, column=4).value = match_formula
ws3.cell(row=res, column=4).font = F_BIG
ws3.cell(row=res, column=4).alignment = Alignment(horizontal="center", vertical="center")
ws3.cell(row=res, column=4).border = THIN
ws3.cell(row=res, column=5, value="мес.").border = THIN
for c in [1, 2, 3, 6, 7]: ws3.cell(row=res, column=c).border = THIN
res += 1

ws3.cell(row=res, column=1, value="Всего внесено:").font = F_BOLD
ws3.merge_cells(start_row=res, start_column=1, end_row=res, end_column=3)
ws3.cell(row=res, column=4, value="$C$2")
ws3.cell(row=res, column=4).number_format = '#,##0.00'
ws3.cell(row=res, column=4).font = F_BOLD; ws3.cell(row=res, column=4).border = THIN
ws3.cell(row=res, column=4).alignment = Alignment(horizontal="center")
for c in [1, 2, 3, 5, 6, 7]: ws3.cell(row=res, column=c).border = THIN
res += 1

calc = res
ws3.cell(row=res, column=1, value="Хочу накопить за N месяцев:").font = F_BOLD
ws3.merge_cells(start_row=res, start_column=1, end_row=res, end_column=3)
ws3.cell(row=res, column=4, value=12).fill = F["input"]
ws3.cell(row=res, column=4).border = THIN
ws3.cell(row=res, column=4).alignment = Alignment(horizontal="center")
ws3.cell(row=res, column=5, value="мес. → нужно").border = THIN
ws3.cell(row=res, column=6, value=f"=PMT($C$4/100/12,D{calc},0,-$C$2)")
ws3.cell(row=res, column=6).number_format = '#,##0.00'
ws3.cell(row=res, column=6).font = F_BOLD; ws3.cell(row=res, column=6).border = THIN
ws3.cell(row=res, column=6).alignment = Alignment(horizontal="center")
ws3.cell(row=res, column=7, value="руб./мес.").border = THIN
for c in [1, 2, 3]: ws3.cell(row=res, column=c).border = THIN
ws3.freeze_panes = "A7"

path = r"C:\Users\win11\Documents\GIT\qasdet.github.io\finance.xlsx"
wb.save(path)
print(f"Done: {path}")
