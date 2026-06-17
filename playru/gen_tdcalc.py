import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ── Стили ──
hdr_font = Font(bold=True, size=11, color="FFFFFF")
hdr_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
sub_fill = PatternFill(start_color="E0E7FF", end_color="E0E7FF", fill_type="solid")
sub_font = Font(bold=True, size=11, color="1E1B4B")
input_fill = PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid")
input_font = Font(bold=True, size=11, color="854D0E")
result_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
result_font = Font(bold=True, size=11, color="166534")
formula_fill = PatternFill(start_color="F3E8FF", end_color="F3E8FF", fill_type="solid")
note_font = Font(size=10, color="6B7280", italic=True)
thin = Side(style='thin', color='C7C7C7')
border = Border(left=thin, right=thin, top=thin, bottom=thin)
wrap = Alignment(wrap_text=True, vertical='center')
center = Alignment(horizontal='center', vertical='center')
num_fmt = '#,##0'

def hdr(ws, r, c, v):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = Alignment(horizontal='center', wrap_text=True); cell.border = border

def sub(ws, r, c, v):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = sub_font; cell.fill = sub_fill; cell.border = border; return cell

def inp(ws, r, c, v):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = input_font; cell.fill = input_fill; cell.border = border; cell.alignment = center
    return cell

def res(ws, r, c, v, fmt=None):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = result_font; cell.fill = result_fill; cell.border = border; cell.alignment = center
    if fmt: cell.number_format = fmt
    return cell

def cell(ws, r, c, v, fmt=None):
    cell = ws.cell(row=r, column=c, value=v)
    cell.border = border; cell.alignment = wrap
    if fmt: cell.number_format = fmt
    return cell

# ═══════════════════════════════════
# ЛИСТ 1: Калькулятор
# ═══════════════════════════════════
ws1 = wb.active
ws1.title = 'Калькулятор'

# Ширины
ws1.column_dimensions['A'].width = 34
ws1.column_dimensions['B'].width = 16
ws1.column_dimensions['C'].width = 18
ws1.column_dimensions['D'].width = 18
ws1.column_dimensions['E'].width = 60
ws1.column_dimensions['F'].width = 20

# Заголовок
ws1.merge_cells('A1:F1')
cell(ws1, 1, 1, 'КАЛЬКУЛЯТОР ТЕСТ-КЕЙСОВ ПО ТЕХНИКАМ ТЕСТ-ДИЗАЙНА')
ws1.row_dimensions[1].height = 30

# Легенда
cell(ws1, 2, 1, '')
cell(ws1, 2, 2, '')
inp(ws1, 2, 2, 'Ввод') if False else None
inp_obj = ws1.cell(row=2, column=2)
inp_obj.value = 'Ввод'
inp_obj.font = input_font; inp_obj.fill = input_fill; inp_obj.border = border; inp_obj.alignment = center
cell(ws1, 2, 3, '')
res(ws1, 2, 3, 'Результат')
cell(ws1, 2, 4, '')
cell(ws1, 2, 5, 'Формула')
inp_obj = ws1.cell(row=2, column=5)
inp_obj.font = formula_fill; inp_obj.fill = formula_fill; inp_obj.border = border

r = 4  # текущая строка

# ── Вспомогательная функция для добавления техники ──
def add_technique(ws, start_r, name, acr, cat, description, inputs, formulas):
    """
    inputs: список (название, ячейка_ввода, значение_по_умолчанию)
    formulas: список (название_результата, формула_Excel_или_текст, формула_текст)
    """
    r = start_r
    # Заголовок техники
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    sub(ws, r, 1, f'{name} [{acr}] — {cat}')
    ws.row_dimensions[r].height = 22
    r += 1
    # Описание
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    cell(ws, r, 1, description)
    ws.row_dimensions[r].height = 28
    r += 1
    # Inputs
    first_input_row = r
    for i, (label, cell_ref, default) in enumerate(inputs):
        cell(ws, r, 1, label)
        inp(ws, r, cell_ref, default)
        ws.row_dimensions[r].height = 22
        r += 1
    # Пустая строка
    r += 1
    # Results header
    hdr(ws, r, 1, 'Показатель')
    hdr(ws, r, 2, 'Min')
    hdr(ws, r, 3, 'Рекомендуется')
    hdr(ws, r, 4, '')
    hdr(ws, r, 5, 'Формула')
    r += 1
    # Results
    for i, (label, formula_min, formula_rec, formula_text) in enumerate(formulas):
        cell(ws, r, 1, label)
        res(ws, r, 2, formula_min, num_fmt)
        res(ws, r, 3, formula_rec, num_fmt)
        cell(ws, r, 4, '')
        cell(ws, r, 5, formula_text)
        ws.row_dimensions[r].height = 22
        r += 1
    # Пустая строка после техники
    r += 1
    return r

# ═══════════════════════════════════
# 1. Эквивалентное разделение (EP)
# ═══════════════════════════════════
r = add_technique(ws1, r,
    'Эквивалентное разделение', 'EP', 'Классические',
    'Разбиение входных данных на классы эквивалентности. Один тест на класс.',
    [('Кол-во классов эквивалентности', 2, 5)],
    [('Тест-кейсов', '=B{}'.format(r+4), '=B{}'.format(r+4) + '+INT(B{}*0.3)'.format(r+4),
      'Min = количество классов; Рек. = классы + 30% на невалидные')]
)
r += 0  # r уже обновлён внутри

# Actually the row numbers are dynamic, so I need to compute them differently.
# Let me restructure this. The issue is that add_technique returns the next r, but the
# formulas need to reference input cells by row. Since rows are dynamic, I'll predefine
# all input row positions and pass them.

# Let me redo the approach: I'll use fixed row positions directly.
wb.remove(ws1)  # remove the empty sheet
wb.create_sheet('Калькулятор', 0)

ws = wb['Калькулятор']

# Column widths
for col, w in {'A': 36, 'B': 16, 'C': 18, 'D': 18, 'E': 65, 'F': 20}.items():
    ws.column_dimensions[col].width = w

# Header
ws.merge_cells('A1:F1')
cell(ws, 1, 1, 'КАЛЬКУЛЯТОР КОЛИЧЕСТВА ТЕСТ-КЕЙСОВ ПО ТЕХНИКАМ ТЕСТ-ДИЗАЙНА')
ws.row_dimensions[1].height = 32

# Legend
cell(ws, 2, 1, '')
inp(ws, 2, 2, 'Ввод')
cell(ws, 2, 3, '')
res(ws, 2, 3, 'Результат')
cell(ws, 2, 4, '')
cell(ws, 2, 5, 'Формула / Примечание')
ws.row_dimensions[2].height = 22

# ═══════════════════════════════════════════════════════════
# ROW MAP — запоминаем строки ввода
# ═══════════════════════════════════════════════════════════
# Каждая техника: 1 строка заголовок, 1 строка описание, N строк ввода,
#   1 пустая, 1 строка заголовков результатов, M строк результатов, 1 пустая

row = 4

def section_title(t):
    global row
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    sub(ws, row, 1, t)
    ws.row_dimensions[row].height = 24
    row += 1

def section_desc(t):
    global row
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    cell(ws, row, 1, t)
    ws.row_dimensions[row].height = 26
    row += 1

def input_row(label, default=0):
    global row
    cell(ws, row, 1, label)
    inp(ws, row, 2, default)
    r = row
    row += 1
    return r

def result_header():
    global row
    hdr(ws, row, 1, 'Показатель')
    hdr(ws, row, 2, 'Min')
    hdr(ws, row, 3, 'Рекомендуется')
    hdr(ws, row, 4, '')
    hdr(ws, row, 5, 'Описание расчёта')
    row += 1

def result_row(label, min_val, rec_val, desc):
    global row
    cell(ws, row, 1, label)
    res(ws, row, 2, min_val, num_fmt)
    res(ws, row, 3, rec_val, num_fmt)
    cell(ws, row, 4, '')
    cell(ws, row, 5, desc)
    ws.row_dimensions[row].height = 22
    row += 1

def blank():
    global row
    row += 1

# ═══════════════════════════════════
# 1. EP
# ═══════════════════════════════════
section_title('1. Эквивалентное разделение [EP] — Классические')
section_desc('Разбиение на классы эквивалентности. Один тест из каждого класса.')
r_ep = input_row('Кол-во классов эквивалентности', 5)
blank()
result_header()
result_row('Тест-кейсов', f'=B{r_ep}', f'=B{r_ep}+INT(B{r_ep}*0.3)',
           'Min = число классов. Рек. = классы + 30% на невалидные классы.')
blank()

# ═══════════════════════════════════
# 2. BVA
# ═══════════════════════════════════
section_title('2. Анализ граничных значений [BVA] — Классические')
section_desc('Тестирование на границах: min, max, и сразу за ними.')
r_bva = input_row('Кол-во границ (диапазонов)', 3)
blank()
result_header()
result_row('2-значное BVA', f'=B{r_bva}*2+1', f'=B{r_bva}*2+1',
           'Min = границы×2 + 1 номинал. Классика: по 2 теста на границу + номинал.')
result_row('3-значное BVA', f'=B{r_bva}*3', f'=B{r_bva}*3',
           'Расширенный: min-1, min, min+1 для каждой границы.')
blank()

# ═══════════════════════════════════
# 3. Pairwise
# ═══════════════════════════════════
section_title('3. Попарное тестирование [Pairwise] — Классические')
section_desc('Все пары комбинаций параметров. ~85% дефектов от пар.')
r_pw_p = input_row('Кол-во параметров', 5)
r_pw_v = input_row('Максимум значений у параметра', 4)
blank()
result_header()
result_row('Приблизительно тест-кейсов', f'=B{r_pw_v}*B{r_pw_v}', f'=B{r_pw_v}*B{r_pw_v}+INT(B{r_pw_p}*0.5)',
           'Грубая оценка: V×V. Точное число даёт PICT/AllPairs.')
result_row('Exhaustive (все комбинации)', f'=B{r_pw_v}^B{r_pw_p}', f'=B{r_pw_v}^B{r_pw_p}',
           'Полный перебор V^P. Для сравнения экономии.')
blank()

# ═══════════════════════════════════
# 4. Decision Table
# ═══════════════════════════════════
section_title('4. Таблица принятия решений [DT] — Классические')
section_desc('Условия → Действия. Полное комбинаторное покрытие логики.')
r_dt = input_row('Кол-во условий (boolean)', 4)
r_dt_rules = input_row('Кол-во правил (если сокращена)', 10)
blank()
result_header()
result_row('Полное покрытие (2^n)', f'=2^B{r_dt}', f'=2^B{r_dt}',
           '2^n, где n = число условий. Каждая строка таблицы = 1 тест.')
result_row('По фактическим правилам', f'=B{r_dt_rules}', f'=B{r_dt_rules}+INT(B{r_dt_rules}*0.1)',
           'Если таблица уже сокращена. Рек. +10% на невалидные комбинации.')
blank()

# ═══════════════════════════════════
# 5. State Transition
# ═══════════════════════════════════
section_title('5. Переходы состояний [STT] — Классические')
section_desc('События → переходы между состояниями. Тестирование конечного автомата.')
r_stt_s = input_row('Кол-во состояний', 5)
r_stt_t = input_row('Кол-во переходов', 12)
blank()
result_header()
result_row('All States', f'=B{r_stt_s}', f'=B{r_stt_s}',
           'По одному тесту на каждое состояние.')
result_row('All Transitions', f'=B{r_stt_t}', f'=B{r_stt_t}+INT(B{r_stt_t}*0.2)',
           'Каждый переход = тест. Рек. +20% на недопустимые переходы.')
result_row('N-Switch (0-switch → 1-switch)', f'=B{r_stt_s}^2', f'=B{r_stt_s}^2+INT(B{r_stt_s}*0.3)',
           'Последовательности из 2 состояний. Для сложных workflow.')
blank()

# ═══════════════════════════════════
# 6. Use Case
# ═══════════════════════════════════
section_title('6. Тестирование сценариев [UCS] — Классические')
section_desc('Сквозные пользовательские сценарии: основной + альтернативные.')
r_uc_m = input_row('Основных сценариев', 2)
r_uc_a = input_row('Альтернативных', 4)
r_uc_e = input_row('Ошибочных', 3)
blank()
result_header()
result_row('Всего сценариев', f'=B{r_uc_m}+B{r_uc_a}+B{r_uc_e}', f'=B{r_uc_m}+B{r_uc_a}+B{r_uc_e}+INT(B{r_uc_a}*0.3)',
           'Min = основной + альт. + ошибочные. Рек. +30% на вариации альтернатив.')
blank()

# ═══════════════════════════════════
# 7. Classification Tree
# ═══════════════════════════════════
section_title('7. Классификация деревьев [CTM] — Формальные')
section_desc('Иерархическая классификация параметров. Комбинации выборов.')
r_ctm_c = input_row('Кол-во классификаций', 4)
r_ctm_v = input_row('Среднее кол-во выборов', 3)
blank()
result_header()
result_row('Все классы', f'=B{r_ctm_v}^B{r_ctm_c}', f'=B{r_ctm_v}^B{r_ctm_c}',
           'V^C — дерево всех комбинаций. Обычно избыточно.')
result_row('Pairwise по классификациям', f'=B{r_ctm_v}*B{r_ctm_v}', f'=B{r_ctm_v}*B{r_ctm_v}+INT(B{r_ctm_c}*1.5)',
           'Как Pairwise: V². Рекомендуемый подход для CTM.')
blank()

# ═══════════════════════════════════
# 8. Orthogonal Arrays
# ═══════════════════════════════════
section_title('8. Ортогональные массивы [OAT] — Формальные')
section_desc('Сбалансированные комбинации (Taguchi). Равномерное покрытие пар.')
r_oat_p = input_row('Кол-во параметров', 4)
r_oat_v = input_row('Кол-во значений', 3)
blank()
result_header()
result_row('Приблизительно тест-кейсов', f'=B{r_oat_v}*B{r_oat_v}', f'=B{r_oat_v}*B{r_oat_v}+INT(B{r_oat_p}*0.3)',
           'Грубо: V². Для точного — таблица L-массива (L9, L16, L25).')
result_row('Exhaustive', f'=B{r_oat_v}^B{r_oat_p}', f'=B{r_oat_v}^B{r_oat_p}',
           'V^P. Сравнение с OAT.')
blank()

# ═══════════════════════════════════
# 9. Code Coverage
# ═══════════════════════════════════
section_title('9. Покрытие кода [CCT] — Формальные')
section_desc('Метрики: строки, ветки, условия, MCDC.')
r_cc_b = input_row('Кол-во веток (branch)', 12)
r_cc_c = input_row('Кол-во условий', 8)
blank()
result_header()
result_row('Branch coverage', f'=B{r_cc_b}', f'=B{r_cc_b}+INT(B{r_cc_b}*0.1)',
           'Каждая ветка = тест. Рек. +10% на граничные.')
result_row('MC/DC покрытие', f'=B{r_cc_c}+1', f'=B{r_cc_c}+INT(B{r_cc_c}*0.3)',
           'Min: C+1. Рек. +30% на сложные условия.')
result_row('Condition coverage', f'=B{r_cc_c}*2', f'=B{r_cc_c}*2',
           'Каждое условие проверяется и true, и false.')
blank()

# ═══════════════════════════════════
# 10. Domain Testing
# ═══════════════════════════════════
section_title('10. Тестирование доменов [DomT] — Формальные')
section_desc('Глубокий анализ одного поля: классы + границы + форматы.')
r_dom_k = input_row('Кол-во классов', 4)
r_dom_b = input_row('Кол-во границ', 6)
blank()
result_header()
result_row('Базовое (EP+BVA)', f'=B{r_dom_k}+INT(B{r_dom_b}*2)', f'=B{r_dom_k}+INT(B{r_dom_b}*2.5)',
           'Min = классы + границы×2. Рек. ×1.25 на спецформаты.')
blank()

# ═══════════════════════════════════
# 11. Cause-Effect
# ═══════════════════════════════════
section_title('11. Причинно-следственный анализ [CEG] — Формальные')
section_desc('Граф причин с логикой AND/OR/NOT → таблица решений.')
r_ceg_c = input_row('Кол-во причин', 5)
r_ceg_e = input_row('Кол-во эффектов', 3)
blank()
result_header()
result_row('Полное покрытие', f'=2^B{r_ceg_c}', f'=2^B{r_ceg_c}',
           '2^причины. Как Decision Table.')
result_row('С учётом эффектов', f'=2^B{r_ceg_c}', f'=2^B{r_ceg_c}+INT(B{r_ceg_e}*0.5)',
           'Те же 2^n + поправка на эффекты.')
blank()

# ═══════════════════════════════════
# 12. Data Flow
# ═══════════════════════════════════
section_title('12. Потоки данных [DFT] — Формальные')
section_desc('Анализ def-use: определение → использование → уничтожение переменных.')
r_dft_v = input_row('Кол-во переменных', 10)
r_dft_d = input_row('Среднее def на переменную', 2)
r_dft_u = input_row('Среднее use на переменную', 3)
blank()
result_header()
result_row('All-defs', f'=B{r_dft_v}*B{r_dft_d}', f'=B{r_dft_v}*B{r_dft_d}',
           'Каждое определение = тест.')
result_row('All-uses', f'=B{r_dft_v}*B{r_dft_d}*B{r_dft_u}', f'=B{r_dft_v}*B{r_dft_d}*B{r_dft_u}',
           'Каждая def-use пара = тест. Наиболее полный.')
blank()

# ═══════════════════════════════════
# 13. Syntax Testing
# ═══════════════════════════════════
section_title('13. Синтаксическое тестирование [SynT] — Формальные')
section_desc('По формальной грамматике (BNF, regex). Генерация по правилам.')
r_synt_r = input_row('Кол-во правил грамматики', 8)
r_synt_c = input_row('Кол-во категорий (валид/невалид)', 4)
blank()
result_header()
result_row('По правилам', f'=B{r_synt_r}', f'=B{r_synt_r}*2',
           'Min = по одному на правило. Рек. = ×2 на невалидные варианты.')
result_row('По категориям', f'=B{r_synt_c}', f'=B{r_synt_c}*2',
           'Min = по категории. Рек. = ×2.')
blank()

# ═══════════════════════════════════
# 14. Risk-Based
# ═══════════════════════════════════
section_title('14. Риск-ориентированное тестирование [RBT] — Эвристические')
section_desc('Приоритизация по риску = вероятность × влияние. Без точной формулы.')
r_rbt_m = input_row('Кол-во модулей', 10)
r_rbt_h = input_row('Из них высокого риска', 3)
r_rbt_md = input_row('Среднего риска', 4)
blank()
result_header()
result_row('Базовый план', f'=B{r_rbt_h}*3+B{r_rbt_md}*2+(B{r_rbt_m}-B{r_rbt_h}-B{r_rbt_md})*1',
           f'=B{r_rbt_h}*5+B{r_rbt_md}*3+(B{r_rbt_m}-B{r_rbt_h}-B{r_rbt_md})*2',
           'Высокий риск ×3–5, средний ×2–3, низкий ×1–2.')
blank()

# ═══════════════════════════════════
# 15. Error Guessing
# ═══════════════════════════════════
section_title('15. Угадывание ошибок [EG] — Эвристические')
section_desc('Опыт и типовые шаблоны. Без формул — оцените количество.')
r_eg = input_row('Оценочное кол-во тест-кейсов', 10)
blank()
result_header()
result_row('Ad-hoc проверок', f'=B{r_eg}', f'=B{r_eg}+5',
           'Min = ваша оценка. Рек. +5 на типовые угрозы: XSS, SQLi, пустой ввод, спецсимволы, переполнение.')
blank()

# ═══════════════════════════════════
# 16. Exploratory
# ═══════════════════════════════════
section_title('16. Исследовательское тестирование [ET] — Эвристические')
section_desc('Чартерная сессия: изучение + тест + анализ.')
r_et_s = input_row('Кол-во чартеров', 5)
r_et_h = input_row('Часов на чартер', 2)
blank()
result_header()
result_row('Оценка по чартерам', f'=B{r_et_s}', f'=B{r_et_s}*2',
           'Min = чартеров. Рек. ×2 на перепроверки.')
blank()

# ═══════════════════════════════════
# 17. Performance
# ═══════════════════════════════════
section_title('17. Тестирование производительности [PerfT] — Специфические')
section_desc('Нагрузка, стресс, стабильность, объём, пик.')
r_perf_s = input_row('Сценариев нагрузки', 3)
r_perf_l = input_row('Уровней нагрузки', 4)
blank()
result_header()
result_row('Всего тестов', f'=B{r_perf_s}*B{r_perf_l}', f'=B{r_perf_s}*(B{r_perf_l}+2)',
           'Min = сценарии × уровни. Рек. +2 уровня (стресс, пик).')
blank()

# ═══════════════════════════════════
# 18. Security
# ═══════════════════════════════════
section_title('18. Тестирование безопасности [SecT] — Специфические')
section_desc('OWASP Top 10, XSS, SQLi, CSRF, аутентификация.')
r_sect_m = input_row('Категорий угроз (OWASP)', 10)
r_sect_t = input_row('Тест-кейсов на категорию', 5)
blank()
result_header()
result_row('Всего тестов', f'=B{r_sect_m}*B{r_sect_t}', f'=B{r_sect_m}*(B{r_sect_t}+2)',
           'Min = категории × тестов. Рек. +2 на категорию (обходные пути).')
blank()

# ═══════════════════════════════════
# 19. Documentation
# ═══════════════════════════════════
section_title('19. Тестирование документации [DocT] — Специфические')
section_desc('Тексты, подписи, сообщения, хелпы, README.')
r_doct_s = input_row('Кол-во экранов/страниц', 15)
r_doct_e = input_row('Элементов на экране', 20)
blank()
result_header()
result_row('Всего элементов', f'=B{r_doct_s}*B{r_doct_e}/10', f'=B{r_doct_s}*B{r_doct_e}/5',
           'Выборочная проверка: ~10–20% элементов.')
blank()

# ═══════════════════════════════════
# 20. Model-Based
# ═══════════════════════════════════
section_title('20. По моделям [MBT] — Формальные')
section_desc('Автогенерация из модели. Зависит от модели.')
r_mbt_s = input_row('Состояний в модели', 6)
r_mbt_t = input_row('Переходов', 15)
blank()
result_header()
result_row('All Transitions', f'=B{r_mbt_t}', f'=B{r_mbt_t}*2',
           'Min = переходов. Рек. ×2 (включая невалидные).')
result_row('N-путей (длина пути)', f'=B{r_mbt_t}*3', f'=B{r_mbt_t}*5',
           'Пути длины 3–5. Рекомендуемый подход для MBT.')
blank()

# ═══════════════════════════════════════════════════════════
# ЛИСТ 2: Пояснения
# ═══════════════════════════════════════════════════════════
ws2 = wb.create_sheet('Пояснения')
ws2.column_dimensions['A'].width = 36
ws2.column_dimensions['B'].width = 80

hdr(ws2, 1, 1, 'Техника')
hdr(ws2, 1, 2, 'Как пользоваться калькулятором')

explanations = [
    ('EP', 'Введите количество выделенных классов эквивалентности (валидные + невалидные). Минимум — по одному тесту на класс. Рекомендуется +30% на невалидные.'),
    ('BVA', 'Введите количество границ (диапазонов). Для поля от 1 до 100 — одна граница. 2-значное BVA: min, max, сразу за границей. 3-значное: min-1, min, min+1.'),
    ('Pairwise', 'Введите число параметров и максимальное количество значений у одного параметра. V² — грубая оценка. Для точного числа используйте PICT (Microsoft) или AllPairs.'),
    ('DT', 'Введите количество булевых условий. 2^n — полный перебор. Если таблица уже сокращена — введите количество правил.'),
    ('STT', 'Состояния — узлы автомата. Переходы — дуги. All States проверяет достижимость. All Transitions — все переходы. N-Switch — последовательности.'),
    ('UCS', 'Основные — happy path. Альтернативные — off-path. Ошибочные — невалидные сценарии. Сумма даёт минимум.'),
    ('CTM', 'Классификации — группы параметров. Выборы — значения в группе. V^C — полное дерево. Pairwise — оптимизация.'),
    ('OAT', 'Укажите параметры и значения. V² — грубая оценка. Выберите L-массив (L9, L16, L25) для точного числа: L9 = 9 тестов для 4×3.'),
    ('CCT', 'Ветки — if/else. Условия — булевы выражения. Branch = каждая ветка. MC/DC = C+1. Condition = true/false для каждого.'),
    ('DomT', 'Глубокий анализ поля: классы эквивалентности + границы + форматы. Базово = EP + BVA × 2.'),
    ('CEG', 'Причины — входные условия. Эффекты — результаты. Полный перебор = 2^причины. Как Decision Table.'),
    ('DFT', 'Переменные — всего. Def — где определяется. Use — где используется. All-uses = V × D × U.'),
    ('SynT', 'Правила — продукционные (BNF). Категории — валид/невалид/спецсимволы. Min — по правилу.'),
    ('RBT', 'Без точной формулы. Min: high×3 + mid×2 + low×1. Рек.: high×5 + mid×3 + low×2.'),
    ('EG', 'Ad-hoc: введите свою оценку. Рекомендуется +5 на типовые уязвимости (XSS, SQLi, пусто, спецсимволы, переполнение).'),
    ('ET', 'Чартеры — цели сессий. Min = чартеров. Рек.: ×2 на перепроверки.'),
    ('PerfT', 'Сценарии — разные профили нагрузки. Уровни — ramp-up, steady, peak, spike.'),
    ('SecT', 'Категории — типы угроз (OWASP категории). Min = категории × тестов.'),
    ('DocT', 'Выборочная проверка: 10–20% элементов. Введите число экранов и примерное число элементов.'),
    ('MBT', 'Состояния и переходы из модели. All Transitions = число переходов. N-пути = переходы × длина пути.'),
]

for r, (tech, desc) in enumerate(explanations, 2):
    cell(ws2, r, 1, tech)
    cell(ws2, r, 2, desc)
    ws2.row_dimensions[r].height = 36

# Сохранение
filepath = r'C:\Users\win11\Documents\GIT\qasdet.github.io\td-calculator.xlsx'
wb.save(filepath)
print(f'Done: {filepath}')
